from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from database import get_db
from security import get_current_user
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import models
import io
from datetime import datetime

router = APIRouter(prefix="/export", tags=["Export"])

BLUE = "4B5FFF"
DARK = "2A0066"
SECONDARY = "5C1FCC"
LIGHT = "F3EEFF"
WHITE = "FFFFFF"
GRAY = "F9F9F9"

STATUS_COLORS = {
    "Postulé": "F3EEFF",
    "Entretien": "FFF0EB",
    "Offre": "EAF3DE",
    "Refusé": "FCEBEB",
    "Sans réponse": "F1EFE8"
}

STATUS_TEXT_COLORS = {
    "Postulé": "5C1FCC",
    "Entretien": "CC3D00",
    "Offre": "3B6D11",
    "Refusé": "A32D2D",
    "Sans réponse": "5F5E5A"
}

def thin_border(color="DDDDDD"):
    side = Side(style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)


@router.get("/excel")
def export_excel(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    applications = db.query(models.Application).filter(
        models.Application.user_id == current_user.id
    ).all()

    wb = Workbook()

    # ══════════════════════════════
    # FEUILLE 1 — Résumé
    # ══════════════════════════════
    cover = wb.active
    cover.title = "Résumé"
    cover.sheet_view.showGridLines = False
    cover.column_dimensions["A"].width = 3
    cover.column_dimensions["B"].width = 28
    cover.column_dimensions["C"].width = 16
    cover.column_dimensions["D"].width = 16

    # Titre
    cover.row_dimensions[2].height = 50
    cell = cover.cell(row=2, column=2, value="JOBBLY")
    cell.font = Font(name="Segoe UI", size=26, bold=True, color=BLUE)
    cell.alignment = Alignment(vertical="center")

    cover.row_dimensions[3].height = 20
    sub = cover.cell(row=3, column=2, value="Tracker de candidatures")
    sub.font = Font(name="Segoe UI", size=12, color=SECONDARY)

    for col in [2, 3, 4]:
        cover.cell(row=4, column=col).fill = PatternFill("solid", fgColor=BLUE)
    cover.row_dimensions[4].height = 3

    cover.row_dimensions[6].height = 18
    cover.cell(row=6, column=2).value = f"Exporté le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
    cover.cell(row=6, column=2).font = Font(name="Segoe UI", size=10, color="888888")

    cover.row_dimensions[7].height = 18
    cover.cell(row=7, column=2).value = f"Utilisateur : {current_user.email}"
    cover.cell(row=7, column=2).font = Font(name="Segoe UI", size=10, color=DARK)

    cover.row_dimensions[8].height = 24
    total_cell = cover.cell(row=8, column=2, value=f"Total : {len(applications)} candidature(s)")
    total_cell.font = Font(name="Segoe UI", size=12, bold=True, color=BLUE)

    cover.row_dimensions[10].height = 18
    cover.cell(row=10, column=2).value = "RÉPARTITION PAR STATUT"
    cover.cell(row=10, column=2).font = Font(name="Segoe UI", size=9, bold=True, color="888888")

    # Header tableau statut
    cover.row_dimensions[11].height = 26
    for col, text in [(2, "Statut"), (3, "Nb"), (4, "%")]:
        cell = cover.cell(row=11, column=col, value=text)
        cell.font = Font(name="Segoe UI", size=10, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(
            horizontal="left" if col == 2 else "center",
            vertical="center",
            indent=1 if col == 2 else 0
        )
        cell.border = thin_border(BLUE)

    status_counts = {}
    for app in applications:
        status_counts[app.status] = status_counts.get(app.status, 0) + 1

    total = len(applications)
    row = 12
    for status, count in status_counts.items():
        cover.row_dimensions[row].height = 22
        bg = STATUS_COLORS.get(status, WHITE)
        txt = STATUS_TEXT_COLORS.get(status, DARK)
        pct = f"{round(count / total * 100)}%" if total > 0 else "0%"

        for col, val, bold in [(2, status, False), (3, count, True), (4, pct, False)]:
            cell = cover.cell(row=row, column=col, value=val)
            cell.font = Font(name="Segoe UI", size=10, bold=bold, color=txt if col == 2 else BLUE if col == 3 else "888888")
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(
                horizontal="left" if col == 2 else "center",
                vertical="center",
                indent=1 if col == 2 else 0
            )
            cell.border = thin_border()
        row += 1

    # Ligne total
    cover.row_dimensions[row].height = 24
    for col, val in [(2, "Total"), (3, total), (4, "100%")]:
        cell = cover.cell(row=row, column=col, value=val)
        cell.font = Font(name="Segoe UI", size=10, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=DARK)
        cell.alignment = Alignment(
            horizontal="left" if col == 2 else "center",
            vertical="center",
            indent=1 if col == 2 else 0
        )
        cell.border = thin_border(DARK)

    # ══════════════════════════════
    # FEUILLE 2 — Candidatures
    # ══════════════════════════════
    ws = wb.create_sheet("Candidatures")
    ws.sheet_view.showGridLines = False

    headers = [
        "Entreprise", "Poste", "Domaine", "Spécialité",
        "Statut", "Ville", "Pays", "Salaire",
        "Date", "Lien", "Notes"
    ]
    col_widths = [22, 28, 20, 30, 14, 14, 14, 14, 14, 35, 40]

    # Header
    ws.row_dimensions[1].height = 32
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(name="Segoe UI", size=11, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border(BLUE)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Données existantes
    for row_idx, app in enumerate(applications, start=2):
        bg = STATUS_COLORS.get(app.status, GRAY if row_idx % 2 == 0 else WHITE)
        txt = STATUS_TEXT_COLORS.get(app.status, DARK)

        row_data = [
            app.company, app.job_title, app.domain or "",
            app.specialty or "", app.status, app.city or "",
            app.country or "", app.salary or "",
            app.date_applied or "", app.link or "", app.notes or ""
        ]

        ws.row_dimensions[row_idx].height = 22
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(
                name="Segoe UI", size=10,
                color=txt if col_idx == 5 else DARK,
                bold=col_idx == 5
            )
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.border = thin_border()
            cell.alignment = Alignment(vertical="center")

    # Lignes vides pour saisie manuelle
    last_data_row = len(applications) + 2
    for row_idx in range(last_data_row, last_data_row + 50):
        bg = GRAY if row_idx % 2 == 0 else WHITE
        ws.row_dimensions[row_idx].height = 22
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.border = thin_border()
            cell.font = Font(name="Segoe UI", size=10, color=DARK)
            cell.alignment = Alignment(vertical="center")

    # Liste déroulante Statut (colonne E)
    dv_status = DataValidation(
        type="list",
        formula1='"Postulé,Entretien,Offre,Refusé,Sans réponse"',
        allow_blank=True,
        showDropDown=False
    )
    ws.add_data_validation(dv_status)
    dv_status.sqref = f"E2:E{last_data_row + 50}"

    # Formatage conditionnel — couleurs selon statut colonne E
    statuses = [
        ("Postulé", "F3EEFF", "5C1FCC"),
        ("Entretien", "FFF0EB", "CC3D00"),
        ("Offre", "EAF3DE", "3B6D11"),
        ("Refusé", "FCEBEB", "A32D2D"),
        ("Sans réponse", "F1EFE8", "5F5E5A"),
    ]

    for status, bg, txt in statuses:
        fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        font = Font(name="Segoe UI", size=10, color=txt, bold=False)
        dxf = DifferentialStyle(fill=fill, font=font)
        rule = Rule(
            type="expression",
            dxf=dxf,
            formula=[f'$E2="{status}"']
        )
        ws.conditional_formatting.add(
            f"A2:{get_column_letter(len(headers))}{last_data_row + 50}",
            rule
        )

    # Freeze + filtres
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # Export
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    date_str = datetime.now().strftime("%Y-%m-%d")
    filename = f"Jobbly_candidatures_{date_str}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )